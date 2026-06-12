"""Extract quantitative results from the ShelterWA accessible housing poll
crosstab workbook into JSON for the dashboard.

Usage: python3 scripts/extract_data.py <path-to-xlsx>
Writes docs/data.js (a JS file assigning the data to window.POLL_DATA so the
dashboard works when opened directly from disk as well as over HTTP).
"""
import json
import sys

import openpyxl

# Columns we expose in the dashboard: (sheet column index, group label)
SEGMENTS = {
    "overall": [(2, "Overall")],
    "age": [(5, "18–34"), (6, "35–54"), (7, "55+")],
    "gender": [(3, "Men"), (4, "Women")],
    "location": [(14, "Metro"), (15, "Regional")],
    "disability": [
        (23, "Personal connection to disability"),
        (25, "Living with a disability"),
        (26, "Caring for someone with a disability"),
        (28, "No connection to disability"),
    ],
}

QUESTIONS = ["SHWA23", "SHWA24", "SHWA25", "SHWA26", "SHWA_D1"]

STANCE_GROUPS = {
    "Strongly support": "support",
    "Somewhat support": "support",
    "Neither support nor oppose": "neutral",
    "Somewhat oppose": "oppose",
    "Strongly oppose": "oppose",
}

# Curated quotes. Each entry references its source sheet row (1-based, header
# = row 1) so every quote is provably drawn from the data. "fragments" are
# verbatim substrings of the response, joined with ellipses when the full
# response is trimmed; None means quote the full response.
WHY_QUOTES = [
    {"row": 122, "fragments": None},
    {"row": 45, "fragments": None},
    {"row": 27, "fragments": None},
    {"row": 399, "fragments": [
        "It makes complete sense and makes society more equitable.",
        "We are a nation with reasonable resources and this is very much a fair concession.",
    ]},
    {"row": 159, "fragments": None},
    {"row": 87, "fragments": None},
    {"row": 349, "fragments": None},
    {"row": 4, "fragments": None},
    {"row": 354, "fragments": [
        "Introducing more and more rules only pushes up the price of housing. Why not let people build to the requirement 'they' want.",
    ]},
]

HOME_QUOTES = [
    {"row": 5, "fragments": None},
    {"row": 61, "fragments": None},
    {"row": 17, "fragments": [
        "Both showers have a step which means my mother if she stays over can not shower in our house.",
        "There are steps in both the front and back of the house",
    ]},
    {"row": 55, "fragments": [
        "I have a house that was built in 1945, the doorways are quite narrow, if I were to use crutches when my knee finally gives out, or end up in a wheelchair from my back a bit earlier than expected, then I could no longer live here.",
        "My niece's unit in Perth is an older type build, my sister can not visit her as the wheelchair doesn't fit.",
    ]},
    {"row": 86, "fragments": [
        "It is not at all accessible has steps at the entry just inside and stairs to the upper rooms, I cannot find another home that is more accessible.",
    ]},
    {"row": 57, "fragments": [
        "I have a mortgage on a home we chose because it required minimal modification to make it suitable for my wheelchair. Wider corridors, wide doorways, and open living spaces.",
        "the standards will also benefit so many others.",
    ]},
    {"row": 13, "fragments": [
        "Any modifications to my grandparents house was done by my grandfather as the need arose. There is now a ramp and rail on the back door",
        "They no longer use the front door due to the steps. My rental is built up off the ground with steeps steps, so my grandparents cannot visit.",
    ]},
    {"row": 219, "fragments": [
        "Wider internal doorways, but no doors. Shower area and toilet are integrated into the same space",
        "The design was to ensure elderly and potentially wheel chair bound people could use the house without restriction due to size or openings. No modifications required.",
    ]},
]

GENDER_LABEL = {"Male": "Man", "Female": "Woman"}


def attribution(row):
    # Raw qual sheets carry gender in col A and age in col B (header labels
    # in the workbook are swapped; values make the order unambiguous).
    gender = GENDER_LABEL.get(row[0], "Person")
    age = row[1]
    location = "Perth metro" if row[2] == "Metro" else "regional WA"
    return f"{gender}, {age}, {location}"


def excerpt(text, fragments):
    text = text.strip()
    if not fragments:
        return text
    for frag in fragments:
        assert frag in text, f"Fragment not found verbatim in source: {frag!r}"
    joined = " … ".join(fragments)
    if not text.startswith(fragments[0]):
        joined = "… " + joined
    if not text.endswith(fragments[-1]):
        joined = joined + " …"
    return joined


def extract_why_themes(ws):
    rows = list(ws.iter_rows(values_only=True))
    counts = {}
    totals = {"all": 0, "support": 0, "neutral": 0, "oppose": 0}
    for row in rows[1:]:
        theme, stance = row[10], row[8]
        if not theme:
            continue
        group = STANCE_GROUPS.get(stance)
        if group is None:
            continue
        totals["all"] += 1
        totals[group] += 1
        c = counts.setdefault(theme, {"all": 0, "support": 0, "neutral": 0, "oppose": 0})
        c["all"] += 1
        c[group] += 1
    themes = [{"label": t, "counts": c} for t, c in
              sorted(counts.items(), key=lambda kv: -kv[1]["all"])]
    return {"totals": totals, "themes": themes}


def extract_why_quotes(ws):
    rows = list(ws.iter_rows(values_only=True))
    quotes = []
    for q in WHY_QUOTES:
        row = rows[q["row"] - 1]
        quotes.append({
            "text": excerpt(row[9], q["fragments"]),
            "stance": row[8],
            "stanceGroup": STANCE_GROUPS[row[8]],
            "theme": row[10],
            "who": attribution(row),
        })
    return quotes


def extract_home_quotes(ws):
    rows = list(ws.iter_rows(values_only=True))
    quotes = []
    for q in HOME_QUOTES:
        row = rows[q["row"] - 1]
        if row[3]:
            connection = "Living with a disability"
        elif row[4]:
            connection = "Cares for someone with a disability"
        else:
            connection = "Close to someone with a disability"
        quotes.append({
            "text": excerpt(row[8], q["fragments"]),
            "connection": connection,
            "who": attribution(row),
        })
    return quotes


def extract_home_themes(ws):
    rows = list(ws.iter_rows(values_only=True))
    themes = []
    in_table = False
    total = None
    for row in rows:
        if row[0] == "Theme / issue":
            in_table = True
            continue
        if in_table:
            if not row[0]:
                break
            count = int(row[1])
            share = float(row[2])
            if total is None:
                total = round(count / share)
            themes.append({"label": row[0], "count": count,
                           "pct": round(share * 100, 1)})
    return {"total": total, "themes": themes}


def extract_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    question_text = (rows[3][0] or "").strip()
    sample_row = rows[3]
    # Response rows: label in col 1, values onward; skip significance rows
    responses = []
    for row in rows[4:]:
        label = row[1]
        if not label:
            continue
        responses.append((label.strip(), row))
    out = {"question": question_text, "segments": {}}
    for seg_name, cols in SEGMENTS.items():
        groups = []
        for col, group_label in cols:
            values = {}
            for label, row in responses:
                v = row[col]
                values[label] = round(float(v) * 100, 1) if v not in (None, "") else None
            groups.append({
                "label": group_label,
                "n": int(sample_row[col]),
                "values": values,
            })
        out["segments"][seg_name] = groups
    return out


def main(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    data = {q: extract_sheet(wb[q]) for q in QUESTIONS}
    data["QUAL_WHY"] = extract_why_themes(wb["SHWA23b"])
    data["QUAL_WHY"]["quotes"] = extract_why_quotes(wb["SHWA23b"])
    data["QUAL_HOME"] = extract_home_themes(wb["SHWA_D1a Qual Summary"])
    data["QUAL_HOME"]["quotes"] = extract_home_quotes(wb["SHWA_D1a"])
    js = "window.POLL_DATA = " + json.dumps(data, indent=2, ensure_ascii=False) + ";\n"
    with open("docs/data.js", "w") as f:
        f.write(js)
    print("Wrote docs/data.js")


if __name__ == "__main__":
    main(sys.argv[1])
