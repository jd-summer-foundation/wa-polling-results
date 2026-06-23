"""Cross-tabulate the talking-point questions (SHWA24/25/26) by each
respondent's initial support level (SHWA23), from the respondent-level raw
data, and write a summary workbook/CSV plus visuals.

Usage:
    python3 scripts/analyze_talking_points_by_support.py <raw_data.xlsx> [outdir]

Hypothesis under test: a respondent's *initial* stance on the standard shapes
how persuasive each talking point is. Note the raw file is unweighted, so
toplines differ from the (weighted) published data tables by ~1pt.
"""
import os
import sys

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl

# --- Coding (verified against the published SHWA23 data table) -------------
SUPPORT_LABELS = {
    1: "Strongly oppose",
    2: "Somewhat oppose",
    3: "Neither support nor oppose",
    4: "Somewhat support",
    5: "Strongly support",
    99: "Unsure",
}
# Display order: most supportive first.
SUPPORT_ORDER = [
    "Strongly support",
    "Somewhat support",
    "Neither support nor oppose",
    "Somewhat oppose",
    "Strongly oppose",
    "Unsure",
]
LIKELY_LABELS = {
    1: "Much less likely",
    2: "Somewhat less likely",
    3: "No difference",
    4: "Somewhat more likely",
    5: "Much more likely",
    99: "Unsure",
}
LIKELY_ORDER = [
    "Much more likely",
    "Somewhat more likely",
    "No difference",
    "Somewhat less likely",
    "Much less likely",
    "Unsure",
]
TALKING_POINTS = {
    "SHWA24": "Occupant benefit / ageing in place",
    "SHWA25": "Cost & housing supply (opposition frame)",
    "SHWA26": "National consistency / WA holdout",
}
COL = {"SHWA23": 14, "SHWA24": 16, "SHWA25": 17, "SHWA26": 18}
SMALL_N = 50  # flag subgroups below this


def load_rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    return rows[1:]  # drop header


def crosstab(rows, tp):
    """Return {support_label: {likely_label: count}} plus group n."""
    table = {s: {l: 0 for l in LIKELY_ORDER} for s in SUPPORT_ORDER}
    for r in rows:
        s = SUPPORT_LABELS.get(r[COL["SHWA23"]])
        l = LIKELY_LABELS.get(r[COL[tp]])
        if s is None or l is None:
            continue
        table[s][l] += 1
    return table


def net_scores(counts, n):
    """More/less/no-diff/net as % of the group (base includes Unsure)."""
    if n == 0:
        return dict(more=0, less=0, nodiff=0, net=0)
    more = counts["Much more likely"] + counts["Somewhat more likely"]
    less = counts["Much less likely"] + counts["Somewhat less likely"]
    return dict(
        more=100 * more / n,
        less=100 * less / n,
        nodiff=100 * counts["No difference"] / n,
        net=100 * (more - less) / n,
    )


# --- Output: CSV + workbook -------------------------------------------------
def write_outputs(rows, outdir):
    tables = {tp: crosstab(rows, tp) for tp in TALKING_POINTS}
    group_n = {s: sum(tables["SHWA24"][s].values()) for s in SUPPORT_ORDER}

    # Tidy long CSV
    csv_path = os.path.join(outdir, "talking_points_by_support.csv")
    with open(csv_path, "w") as f:
        f.write("talking_point,talking_point_label,initial_support,group_n,"
                "low_base_flag,response,count,pct_of_group\n")
        for tp, label in TALKING_POINTS.items():
            for s in SUPPORT_ORDER:
                n = group_n[s]
                flag = "y" if n < SMALL_N else ""
                for l in LIKELY_ORDER:
                    c = tables[tp][s][l]
                    pct = round(100 * c / n, 1) if n else ""
                    f.write(f'{tp},"{label}","{s}",{n},{flag},"{l}",{c},{pct}\n')

    # Workbook
    wb = openpyxl.Workbook()
    _sheet_readme(wb.active, group_n)
    for tp, label in TALKING_POINTS.items():
        _sheet_crosstab(wb.create_sheet(f"{tp} %"), tp, label, tables[tp], group_n, pct=True)
        _sheet_crosstab(wb.create_sheet(f"{tp} counts"), tp, label, tables[tp], group_n, pct=False)
    _sheet_netscores(wb.create_sheet("Net scores"), tables, group_n)
    xlsx_path = os.path.join(outdir, "talking_points_by_support.xlsx")
    wb.save(xlsx_path)
    return tables, group_n, csv_path, xlsx_path


def _sheet_readme(ws, group_n):
    ws.title = "README"
    lines = [
        ["Talking points by initial support level — WA Accessible Housing Poll (June 2026)"],
        ["Source: 26305 ShelterWA Accessible Housing Poll raw data (respondent level), n=762."],
        ["Unweighted; toplines differ from the weighted published data tables by ~1 percentage point."],
        [""],
        ["Hypothesis: a respondent's initial stance (SHWA23) shapes how each talking point lands."],
        [""],
        ["SHWA23 asked initial support for WA adopting the liveable housing design standard."],
        ["SHWA24/25/26 each presented an argument, then asked: more or less likely to support?"],
        ["  SHWA24 = " + TALKING_POINTS["SHWA24"]],
        ["  SHWA25 = " + TALKING_POINTS["SHWA25"]],
        ["  SHWA26 = " + TALKING_POINTS["SHWA26"]],
        [""],
        ["Percentages are column % within each initial-support group (base includes 'Unsure')."],
        ["NET more = much+somewhat more likely; NET less = much+somewhat less likely; Net = more - less."],
        [f"Subgroups with n<{SMALL_N} are flagged and should be read as directional only."],
        [""],
        ["Initial support group sizes (n):"],
    ]
    for row in lines:
        ws.append(row)
    for s in SUPPORT_ORDER:
        flag = "  (low base, n<50)" if group_n[s] < SMALL_N else ""
        ws.append([f"  {s}", group_n[s], flag])
    ws.column_dimensions["A"].width = 80


def _sheet_crosstab(ws, tp, label, table, group_n, pct):
    ws.append([f"{tp}: {label}"])
    ws.append(["Column % within initial-support group" if pct else "Counts"])
    ws.append([])
    header = ["Response \\ Initial support"] + [
        f"{s} (n={group_n[s]})" for s in SUPPORT_ORDER]
    ws.append(header)
    for l in LIKELY_ORDER:
        row = [l]
        for s in SUPPORT_ORDER:
            n = group_n[s]
            c = table[s][l]
            row.append(round(100 * c / n, 1) if (pct and n) else c)
        ws.append(row)
    # Net rows
    ws.append([])
    for key, name in [("more", "NET more likely"), ("less", "NET less likely"),
                      ("net", "Net (more - less)")]:
        row = [name]
        for s in SUPPORT_ORDER:
            ns = net_scores(table[s], group_n[s])
            row.append(round(ns[key], 1) if pct else "")
        ws.append(row)
    ws.column_dimensions["A"].width = 28


def _sheet_netscores(ws, tables, group_n):
    ws.append(["Net score (% more likely minus % less likely), by initial support"])
    ws.append([])
    ws.append(["Initial support", "n", "low base"] +
              [f"{tp}: {lbl}" for tp, lbl in TALKING_POINTS.items()])
    for s in SUPPORT_ORDER:
        flag = "y" if group_n[s] < SMALL_N else ""
        row = [s, group_n[s], flag]
        for tp in TALKING_POINTS:
            row.append(round(net_scores(tables[tp][s], group_n[s])["net"], 1))
        ws.append(row)
    ws.column_dimensions["A"].width = 28
    for col in "DEF":
        ws.column_dimensions[col].width = 30


# --- Visuals ----------------------------------------------------------------
SUPPORT_SHORT = {
    "Strongly support": "Strongly\nsupport",
    "Somewhat support": "Somewhat\nsupport",
    "Neither support nor oppose": "Neither",
    "Somewhat oppose": "Somewhat\noppose",
    "Strongly oppose": "Strongly\noppose",
    "Unsure": "Unsure",
}
# Diverging palette: more-likely greens, no-diff grey, less-likely reds.
RESP_COLOR = {
    "Much more likely": "#1a7a3a",
    "Somewhat more likely": "#79c879",
    "No difference": "#c9c9c9",
    "Somewhat less likely": "#e8927c",
    "Much less likely": "#b2182b",
    "Unsure": "#efe6c8",
}
TP_COLOR = {"SHWA24": "#2c7fb8", "SHWA25": "#d95f0e", "SHWA26": "#756bb1"}


def _groups_for_chart(group_n):
    """Plot groups with a usable base; flag the small ones in labels."""
    return [s for s in SUPPORT_ORDER if group_n[s] > 0]


def chart_diverging(tables, group_n, outdir):
    groups = _groups_for_chart(group_n)
    fig, axes = plt.subplots(1, 3, figsize=(15, 6), sharey=True)
    order_more = ["Much more likely", "Somewhat more likely"]
    order_less = ["Somewhat less likely", "Much less likely"]
    y = np.arange(len(groups))[::-1]  # strong support at top
    for ax, (tp, label) in zip(axes, TALKING_POINTS.items()):
        for yi, s in zip(y, groups):
            n = group_n[s]
            pct = {l: 100 * tables[tp][s][l] / n for l in LIKELY_ORDER}
            # more likely to the right of zero
            left = 0
            for l in order_more:
                ax.barh(yi, pct[l], left=left, color=RESP_COLOR[l],
                        edgecolor="white", height=0.72)
                left += pct[l]
            # less likely to the left of zero
            right = 0
            for l in order_less:
                ax.barh(yi, -pct[l], left=right, color=RESP_COLOR[l],
                        edgecolor="white", height=0.72)
                right -= pct[l]
            # net marker
            net = net_scores(tables[tp][s], n)["net"]
            ax.plot(net, yi, "D", color="black", markersize=5, zorder=5)
        ax.axvline(0, color="#444", lw=0.8)
        ax.set_title(f"{tp}\n{label}", fontsize=10)
        ax.set_xlim(-60, 100)
        ax.set_xlabel("← less likely    % of group    more likely →", fontsize=8)
        ax.grid(axis="x", color="#eee", zorder=0)
    ylabels = [f"{SUPPORT_SHORT[s]}\n(n={group_n[s]})" +
               ("*" if group_n[s] < SMALL_N else "") for s in groups][::-1]
    axes[0].set_yticks(y[::-1])
    axes[0].set_yticklabels(ylabels, fontsize=8)
    handles = [plt.Rectangle((0, 0), 1, 1, color=RESP_COLOR[l]) for l in
               ["Much more likely", "Somewhat more likely", "Somewhat less likely",
                "Much less likely"]]
    handles.append(plt.Line2D([0], [0], marker="D", color="black", lw=0,
                              markersize=6))
    labels = ["Much more likely", "Somewhat more likely", "Somewhat less likely",
              "Much less likely", "Net (more − less)"]
    fig.legend(handles, labels, loc="lower center", ncol=5, fontsize=8,
               frameon=False, bbox_to_anchor=(0.5, -0.02))
    fig.suptitle("How each talking point shifts support, by initial support level",
                 fontsize=13, fontweight="bold")
    fig.text(0.5, 0.04, "'No difference' and 'Unsure' omitted from bars; "
             "* n<50, directional only. Unweighted n=762.",
             ha="center", fontsize=8, color="#666")
    fig.tight_layout(rect=[0, 0.07, 1, 0.95])
    p = os.path.join(outdir, "fig1_diverging_by_support.png")
    fig.savefig(p, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return p


def chart_netscore(tables, group_n, outdir):
    groups = _groups_for_chart(group_n)
    x = np.arange(len(groups))
    w = 0.26
    fig, ax = plt.subplots(figsize=(11, 6))
    for i, (tp, label) in enumerate(TALKING_POINTS.items()):
        vals = [net_scores(tables[tp][s], group_n[s])["net"] for s in groups]
        bars = ax.bar(x + (i - 1) * w, vals, w, label=f"{tp}: {label}",
                      color=TP_COLOR[tp])
        for b, v in zip(bars, vals):
            ax.annotate(f"{v:.0f}", (b.get_x() + b.get_width() / 2,
                        v + (1.5 if v >= 0 else -3)), ha="center", fontsize=7)
    ax.axhline(0, color="#444", lw=0.8)
    labels = [f"{SUPPORT_SHORT[s]}\n(n={group_n[s]})" +
              ("*" if group_n[s] < SMALL_N else "") for s in groups]
    ax.set_xticks(x)
    ax.set_xticklabels(labels, fontsize=9)
    ax.set_ylabel("Net score: % more likely − % less likely")
    ax.set_title("Net persuasion of each talking point, by initial support level",
                 fontsize=13, fontweight="bold")
    ax.legend(fontsize=8, frameon=False)
    ax.grid(axis="y", color="#eee")
    fig.text(0.5, 0.005, "* n<50, directional only. Unweighted n=762.",
             ha="center", fontsize=8, color="#666")
    fig.tight_layout(rect=[0, 0.03, 1, 1])
    p = os.path.join(outdir, "fig2_netscore_by_support.png")
    fig.savefig(p, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return p


def chart_heatmap(tables, group_n, outdir):
    groups = _groups_for_chart(group_n)
    tps = list(TALKING_POINTS)
    mat = np.array([[net_scores(tables[tp][s], group_n[s])["more"]
                     for tp in tps] for s in groups])
    import textwrap
    fig, ax = plt.subplots(figsize=(8.5, 6.2))
    im = ax.imshow(mat, cmap="Greens", vmin=0, vmax=100, aspect="auto")
    ax.set_xticks(range(len(tps)))
    ax.set_xticklabels(
        [f"{tp}\n" + "\n".join(textwrap.wrap(TALKING_POINTS[tp], 20)) for tp in tps],
        fontsize=8)
    ax.set_yticks(range(len(groups)))
    ax.set_yticklabels([f"{s} (n={group_n[s]})" +
                        ("*" if group_n[s] < SMALL_N else "") for s in groups],
                       fontsize=9)
    for i in range(len(groups)):
        for j in range(len(tps)):
            v = mat[i, j]
            ax.text(j, i, f"{v:.0f}%", ha="center", va="center",
                    color="white" if v > 55 else "black", fontsize=10)
    ax.set_title("% 'more likely to support' (NET) by initial support level",
                 fontsize=12, fontweight="bold")
    fig.colorbar(im, ax=ax, label="% more likely", shrink=0.8)
    fig.text(0.5, 0.005, "* n<50, directional only. Unweighted n=762.",
             ha="center", fontsize=8, color="#666")
    fig.tight_layout(rect=[0, 0.06, 1, 1])
    p = os.path.join(outdir, "fig3_heatmap_more_likely.png")
    fig.savefig(p, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return p


def main(path, outdir="analysis"):
    os.makedirs(outdir, exist_ok=True)
    rows = load_rows(path)
    tables, group_n, csv_path, xlsx_path = write_outputs(rows, outdir)
    p1 = chart_diverging(tables, group_n, outdir)
    p2 = chart_netscore(tables, group_n, outdir)
    p3 = chart_heatmap(tables, group_n, outdir)
    print("Wrote:")
    for p in (csv_path, xlsx_path, p1, p2, p3):
        print("  ", p)
    # Console summary of net scores
    print("\nNet score (% more − % less likely):")
    print("  %-28s %8s %8s %8s" % ("Initial support", *TALKING_POINTS))
    for s in SUPPORT_ORDER:
        nets = [net_scores(tables[tp][s], group_n[s])["net"] for tp in TALKING_POINTS]
        print("  %-28s %7.1f%% %7.1f%% %7.1f%%  (n=%d)" %
              (s, nets[0], nets[1], nets[2], group_n[s]))


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else "analysis")
